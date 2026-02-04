# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
import pyodbc
import traceback
import datetime
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from tkcalendar import DateEntry  # 需要安装: pip install tkcalendar

# ============== 用户配置区 ==============
DB_CONN_STRING = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=192.168.0.117;"
    "DATABASE=FQD;"
    "UID=zhitan;"
    "PWD=Zt@forcome;"
)

# 截图中的关键配置
ROW_IDX_HEADER_MAIN = 2  # 主表头所在行
ROW_IDX_HEADER_DATE = 3  # 日期表头所在行
ROW_IDX_DATA_START = 4  # 数据起始行

COL_NAME_WORKSHOP = "车间"
COL_NAME_WO_TYPE = "单别"
COL_NAME_WO_NO = "工单单号"

# 设定要保留的列索引 (1-based, 对应 Excel 的 A=1, B=2...)
# B=2, T=20
# 保留: B(2)-H(8), J(10)-P(16), T(20)
# 排除: I(9), Q(17), R(18), S(19) 以及 A(1)
KEEP_COL_INDICES = [2, 3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 14, 15, 16, 20]


# ============== 应用程序类 ==============

class DailyPlanAvailabilityApp:
    def __init__(self, root):
        self.root = root
        self.root.title("每日排程齐套分析工具 v5.0 (新文件生成版)")
        self.root.geometry("1000x700")

        # 样式定义
        self.red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        self.green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        self.header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))

        # 变量绑定
        self.file_path = tk.StringVar()
        self.sheet_name = tk.StringVar()
        self.selected_workshop = tk.StringVar()

        # 日期范围控制
        self.is_date_range = tk.BooleanVar(value=False)

        # 缓存数据
        self.date_column_map = {}
        self.col_map_main = {}
        self.header_names_map = {}  # 存储保留列的表头名称 {col_idx: name}

        self._create_widgets()

    def _create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 1. 文件选择
        file_frame = ttk.LabelFrame(main_frame, text="1. 数据源", padding="5")
        file_frame.pack(fill=tk.X, pady=5)

        ttk.Entry(file_frame, textvariable=self.file_path, width=50).pack(side=tk.LEFT, padx=5)
        ttk.Button(file_frame, text="浏览Excel...", command=self._select_file).pack(side=tk.LEFT, padx=5)

        ttk.Label(file_frame, text="   工作表:").pack(side=tk.LEFT)
        self.sheet_combo = ttk.Combobox(file_frame, textvariable=self.sheet_name, state="disabled", width=15)
        self.sheet_combo.pack(side=tk.LEFT, padx=5)
        self.sheet_combo.bind("<<ComboboxSelected>>", self._on_sheet_selected)

        # 2. 筛选设置
        filter_frame = ttk.LabelFrame(main_frame, text="2. 计划筛选", padding="10")
        filter_frame.pack(fill=tk.X, pady=5)

        # --- 日期选择区域 ---
        date_frame = ttk.Frame(filter_frame)
        date_frame.pack(side=tk.LEFT, fill=tk.X)

        # 勾选框
        ttk.Checkbutton(date_frame, text="选择日期范围", variable=self.is_date_range, command=self._toggle_date_mode).pack(
            side=tk.LEFT, padx=(0, 10))

        # 开始日期
        ttk.Label(date_frame, text="开始日期:").pack(side=tk.LEFT)
        self.date_start = DateEntry(date_frame, width=12, background='darkblue', foreground='white', borderwidth=2,
                                    date_pattern='yyyy/mm/dd')
        self.date_start.pack(side=tk.LEFT, padx=5)

        # 结束日期 (初始可能隐藏或禁用，这里用Pack动态控制比较麻烦，直接一直显示但根据状态判断)
        self.lbl_end = ttk.Label(date_frame, text="结束日期:")
        self.date_end = DateEntry(date_frame, width=12, background='darkblue', foreground='white', borderwidth=2,
                                  date_pattern='yyyy/mm/dd')

        # 初始状态更新
        self._toggle_date_mode()

        # 车间选择
        ttk.Label(filter_frame, text="选择车间:").pack(side=tk.LEFT, padx=(30, 5))
        self.workshop_combo = ttk.Combobox(filter_frame, textvariable=self.selected_workshop, state="disabled",
                                           width=20)
        self.workshop_combo.pack(side=tk.LEFT, padx=5)

        # 3. 操作区
        action_frame = ttk.LabelFrame(main_frame, text="3. 执行", padding="10")
        action_frame.pack(fill=tk.X, pady=10)

        btn = ttk.Button(action_frame, text="生成缺料分析文件 (另存为)", command=self._run_analysis_batch)
        btn.pack(fill=tk.X, padx=100)

        # 4. 日志
        self.log_text = tk.Text(main_frame, height=15, state="disabled", font=("Consolas", 9), bg="#F0F0F0")
        self.log_text.pack(fill=tk.BOTH, expand=True, pady=5)

    def _toggle_date_mode(self):
        """切换日期选择模式"""
        if self.is_date_range.get():
            self.lbl_end.pack(side=tk.LEFT, padx=(10, 0))
            self.date_end.pack(side=tk.LEFT, padx=5)
        else:
            self.lbl_end.pack_forget()
            self.date_end.pack_forget()

    def _log(self, msg):
        self.log_text.config(state="normal")
        self.log_text.insert(tk.END, f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state="disabled")
        self.root.update_idletasks()

    def _select_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if path:
            self.file_path.set(path)
            try:
                wb = openpyxl.load_workbook(path, read_only=True)
                self.sheet_combo['values'] = wb.sheetnames
                if wb.sheetnames:
                    self.sheet_combo.current(0)
                    self._on_sheet_selected(None)
                self.sheet_combo.config(state="readonly")
            except Exception as e:
                messagebox.showerror("错误", f"无法打开文件: {e}")

    def _on_sheet_selected(self, event):
        file_path = self.file_path.get()
        sheet_name = self.sheet_name.get()
        if not file_path or not sheet_name: return

        self._log("正在扫描Excel结构...")
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[sheet_name]

            # 1. 扫描主表头 (第2行) 并记录需要保留的列名
            self.col_map_main = {}
            self.header_names_map = {}

            # 遍历第2行获取所有列头
            for idx, cell in enumerate(ws[ROW_IDX_HEADER_MAIN], start=1):
                val = str(cell.value).strip() if cell.value else ""
                if val:
                    self.col_map_main[val] = idx

                # 如果这一列在我们要保留的列表中，记录它的名字
                if idx in KEEP_COL_INDICES:
                    self.header_names_map[idx] = val

            # 检查关键列
            required_cols = [COL_NAME_WORKSHOP, COL_NAME_WO_TYPE, COL_NAME_WO_NO]
            missing = [c for c in required_cols if c not in self.col_map_main]
            if missing:
                messagebox.showwarning("警告", f"未找到关键列: {missing}")
                return

            # 2. 扫描日期列 (第3行)
            self.date_column_map = {}
            for cell in ws[ROW_IDX_HEADER_DATE]:
                val = cell.value
                date_obj = self._parse_excel_date(val)
                if date_obj:
                    # 统一转为 datetime.date 对象作为 Key
                    self.date_column_map[date_obj] = cell.column

            date_keys = sorted(list(self.date_column_map.keys()))
            if not date_keys:
                self._log("警告: 在第3行未找到任何日期格式的表头！")
            else:
                self._log(f"Excel中检测到排程日期范围: {date_keys[0]} 至 {date_keys[-1]}")

            # 3. 扫描车间
            col_ws_idx = self.col_map_main[COL_NAME_WORKSHOP]
            workshops = set()
            for row in ws.iter_rows(min_row=ROW_IDX_DATA_START, min_col=col_ws_idx, max_col=col_ws_idx,
                                    values_only=True):
                if row[0]: workshops.add(str(row[0]).strip())

            self.workshop_combo['values'] = ["全部车间"] + sorted(list(workshops))
            self.workshop_combo.current(0)
            self.workshop_combo.config(state="readonly")

        except Exception as e:
            traceback.print_exc()
            self._log(f"扫描失败: {e}")

    def _parse_excel_date(self, val):
        """解析为 datetime.date 对象"""
        if val is None: return None
        try:
            dt = None
            if isinstance(val, datetime.datetime):
                dt = val.date()
            elif isinstance(val, datetime.date):
                dt = val
            elif isinstance(val, (int, float)):
                dt = (datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(val))).date()
            elif isinstance(val, str):
                # 尝试解析文本
                parts = val.strip().split('/')
                if len(parts) == 2:  # M/D 格式，假设为当前年或2026(根据截图)
                    # 这里为了匹配截图中的年份，如果解析不到年份，需要结合上下文。
                    # 简化处理：尝试用当前年份拼凑，如果不对，建议Excel里用标准日期格式
                    now_year = datetime.datetime.now().year
                    # 截图里是2026年，这里也可以硬编码尝试解析，或者就依靠Excel本身如果是日期格式最好
                    # 暂且返回None让用户去检查Excel格式，或者尝试解析
                    return None
                elif len(parts) == 3:
                    dt = datetime.datetime.strptime(val.strip(), "%Y/%m/%d").date()
            return dt
        except:
            return None

    def _get_target_dates(self):
        """获取用户选择的日期列表"""
        start_date = self.date_start.get_date()  # 返回 datetime.date

        if self.is_date_range.get():
            end_date = self.date_end.get_date()
            if end_date < start_date:
                messagebox.showerror("日期错误", "结束日期不能早于开始日期！")
                return []
        else:
            end_date = start_date

        target_dates = []
        curr = start_date
        while curr <= end_date:
            target_dates.append(curr)
            curr += datetime.timedelta(days=1)

        return target_dates

    def _run_analysis_batch(self):
        target_dates = self._get_target_dates()
        if not target_dates: return

        file_path = self.file_path.get()
        sheet_name = self.sheet_name.get()
        target_workshop = self.selected_workshop.get()

        # 验证Excel中是否存在这些日期
        valid_dates = []
        for d in target_dates:
            if d in self.date_column_map:
                valid_dates.append(d)
            else:
                self._log(f"跳过: Excel中未找到日期列 {d}")

        if not valid_dates:
            messagebox.showwarning("无有效日期", "所选日期在Excel表头中均未找到对应列。")
            return

        # 确定文件名
        if len(valid_dates) == 1:
            date_str = valid_dates[0].strftime("%Y-%m-%d")  # 替换斜杠以免文件名非法
            default_name = f"{date_str}缺料分析.xlsx"
        else:
            start_s = valid_dates[0].strftime("%Y-%m-%d")
            end_s = valid_dates[-1].strftime("%Y-%m-%d")
            default_name = f"{start_s}至{end_s}缺料分析.xlsx"

        save_path = filedialog.asksaveasfilename(
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not save_path: return

        try:
            self._log("=" * 50)
            self._log(f"开始批量分析... 共 {len(valid_dates)} 天")

            # 创建新工作簿
            new_wb = openpyxl.Workbook()
            # 删除默认Sheet
            if "Sheet" in new_wb.sheetnames:
                del new_wb["Sheet"]

            # 遍历日期处理
            for d in valid_dates:
                sheet_title = d.strftime("%Y-%m-%d")  # 使用日期作为Sheet名
                self._log(f"正在处理: {sheet_title}")

                # 1. 提取基础数据
                col_idx = self.date_column_map[d]
                plans_data = self._extract_data_for_date(file_path, sheet_name, col_idx, target_workshop)

                if not plans_data:
                    self._log(f"  -> {sheet_title} 无排产数据")
                    # 即使没数据也创建一个空Sheet
                    new_ws = new_wb.create_sheet(title=sheet_title)
                    self._write_headers(new_ws)
                    continue

                # 2. ERP 查询与计算
                wo_keys = list(set(p['wo_key'] for p in plans_data))
                wo_details = self._fetch_erp_data(wo_keys)

                all_parts = set()
                for w in wo_details.values():
                    for b in w['bom']: all_parts.add(b['part'])

                inventory = self._fetch_inventory(list(all_parts))

                # 3. 计算逻辑
                results = self._simulate(plans_data, wo_details, inventory)

                # 4. 写入新Sheet
                new_ws = new_wb.create_sheet(title=sheet_title)
                self._write_new_sheet(new_ws, plans_data, results)
                self._log(f"  -> {sheet_title} 处理完成")

            new_wb.save(save_path)
            messagebox.showinfo("完成", f"文件已保存至:\n{save_path}")
            self._log("全部完成。")

        except Exception as e:
            traceback.print_exc()
            self._log(f"错误: {e}")
            messagebox.showerror("运行错误", str(e))

    def _extract_data_for_date(self, file_path, sheet_name, date_col_idx, filter_ws):
        """
        读取原文件，提取指定列的数据。
        注意：这里无论是否筛选，iter_rows默认读取所有行。
        这符合'另存一份原文件去除筛选'的逻辑效果（直接读取底层数据）。
        """
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]

        c_ws = self.col_map_main[COL_NAME_WORKSHOP]
        c_type = self.col_map_main[COL_NAME_WO_TYPE]
        c_no = self.col_map_main[COL_NAME_WO_NO]

        extracted_rows = []

        for row in ws.iter_rows(min_row=ROW_IDX_DATA_START):
            try:
                # 检查日期列是否有排产
                if date_col_idx > len(row): continue
                daily_qty = row[date_col_idx - 1].value  # 0-based tuple

                if isinstance(daily_qty, (int, float)) and daily_qty > 0:
                    # 检查车间
                    curr_ws = row[c_ws - 1].value
                    curr_ws = str(curr_ws).strip() if curr_ws else "未分类"

                    if filter_ws != "全部车间" and curr_ws != filter_ws:
                        continue

                    # 读取该行的基础信息（保留列）
                    row_data = {}
                    for target_col_idx in KEEP_COL_INDICES:
                        if target_col_idx <= len(row):
                            val = row[target_col_idx - 1].value
                            row_data[target_col_idx] = val
                        else:
                            row_data[target_col_idx] = None

                    # 关键Key
                    wo_type = row[c_type - 1].value
                    wo_no = row[c_no - 1].value

                    if wo_type and wo_no:
                        extracted_rows.append({
                            'base_data': row_data,  # 字典 {col_idx: value}
                            'wo_key': (str(wo_type).strip(), str(wo_no).strip()),
                            'daily_qty': float(daily_qty),
                            'workshop': curr_ws
                        })
            except IndexError:
                continue
        return extracted_rows

    def _fetch_erp_data(self, wo_keys):
        # ... (保持原有逻辑不变)
        if not wo_keys: return {}
        conditions = []
        for t, n in wo_keys:
            conditions.append(f"(TA.TA001='{t}' AND TA.TA002='{n}')")

        if not conditions: return {}

        # 分批处理防止SQL过长 (简单分批，每批200个)
        batch_size = 200
        all_data = defaultdict(lambda: {'total': 0, 'bom': []})

        for i in range(0, len(conditions), batch_size):
            batch_conds = conditions[i:i + batch_size]
            where_sql = " OR ".join(batch_conds)

            sql = f"""
                SELECT 
                    RTRIM(TA.TA001) as ta001, RTRIM(TA.TA002) as ta002, 
                    TA.TA015 as wo_total_qty,
                    RTRIM(TB.TB003) as part_no, ISNULL(RTRIM(MB.MB002),'') as part_name,
                    TB.TB004 as req_qty, TB.TB005 as iss_qty
                FROM MOCTA TA
                INNER JOIN MOCTB TB ON TA.TA001 = TB.TB001 AND TA.TA002 = TB.TB002
                LEFT JOIN INVMB MB ON TB.TB003 = MB.MB001
                WHERE {where_sql}
            """
            try:
                with pyodbc.connect(DB_CONN_STRING) as conn:
                    df = pd.read_sql(sql, conn)
                    for _, row in df.iterrows():
                        k = (row['ta001'], row['ta002'])
                        all_data[k]['total'] = float(row['wo_total_qty'])
                        all_data[k]['bom'].append({
                            'part': row['part_no'],
                            'name': row['part_name'],
                            'req': float(row['req_qty']),
                            'iss': float(row['iss_qty'])
                        })
            except Exception as e:
                self._log(f"数据库查询批次失败: {e}")

        return all_data

    def _fetch_inventory(self, parts):
        # ... (保持原有逻辑不变)
        if not parts: return {}
        # 简单处理，如果parts太多可能报错
        # 同样建议分批
        unique_parts = list(set(parts))
        inventory = {}
        batch_size = 500

        for i in range(0, len(unique_parts), batch_size):
            batch_parts = unique_parts[i:i + batch_size]
            p_str = ",".join(f"'{p}'" for p in batch_parts)
            sql = f"SELECT RTRIM(MC001) as p, SUM(MC007) as q FROM INVMC WHERE MC001 IN ({p_str}) GROUP BY MC001"
            try:
                with pyodbc.connect(DB_CONN_STRING) as conn:
                    df = pd.read_sql(sql, conn)
                inventory.update(pd.Series(df.q.values, index=df.p).to_dict())
            except:
                pass
        return inventory

    def _simulate(self, plans_data, wo_data, inventory):
        """
        计算逻辑保持不变
        Returns: list of dicts with result info appended
        """
        running_inv = inventory.copy()
        calculated_results = []

        for p in plans_data:
            wo_key = p['wo_key']
            daily_qty = p['daily_qty']

            info = wo_data.get(wo_key)

            res_item = {
                'rate': 0.0,
                'achievable': 0,
                'shortage_str': "",
                'is_short': False
            }

            if not info or not info['bom']:
                res_item['shortage_str'] = "无ERP信息"
                res_item['is_short'] = True
                calculated_results.append(res_item)
                continue

            wo_total_qty = info['total']

            items_needed = 0
            items_kitted = 0
            shortage_details = []
            to_deduct = {}
            is_fully_kitted = True
            min_possible_sets = 9999999

            for bom in info['bom']:
                part = bom['part']
                remain_issue = max(0, bom['req'] - bom['iss'])
                unit_usage = bom['req'] / wo_total_qty if wo_total_qty > 0 else 0
                theo_demand = daily_qty * unit_usage
                net_demand = min(remain_issue, theo_demand)

                if net_demand <= 0.0001: continue

                items_needed += 1
                to_deduct[part] = net_demand
                current_stock = running_inv.get(part, 0)

                can_make = int(current_stock // unit_usage) if unit_usage > 0 else 999999
                min_possible_sets = min(min_possible_sets, can_make)

                if current_stock >= net_demand - 0.0001:  # 浮点容差
                    items_kitted += 1
                else:
                    is_fully_kitted = False
                    short_qty = net_demand - current_stock
                    shortage_details.append(f"{part} {bom['name']}(缺{short_qty:.1f})")

            actual_possible_sets = min(int(daily_qty), min_possible_sets)
            if items_needed == 0:
                actual_possible_sets = int(daily_qty)
                is_fully_kitted = True

            rate = (items_kitted / items_needed) if items_needed > 0 else 1.0

            res_item['rate'] = rate
            res_item['achievable'] = actual_possible_sets
            res_item['is_short'] = not is_fully_kitted
            if shortage_details:
                res_item['shortage_str'] = ",".join(shortage_details)

            calculated_results.append(res_item)

            # 库存扣减 (All-or-Nothing)
            if is_fully_kitted:
                for part, qty in to_deduct.items():
                    running_inv[part] -= qty

        return calculated_results

    def _write_headers(self, ws):
        """写入新表头"""
        # 写入原有保留列的表头
        current_col = 1
        for idx in KEEP_COL_INDICES:
            cell = ws.cell(row=1, column=current_col)
            cell.value = self.header_names_map.get(idx, "")
            cell.fill = self.header_fill
            cell.border = self.thin_border
            current_col += 1

        # 写入新列
        new_headers = ["齐套率", "可产数量", "缺料信息"]
        for h in new_headers:
            cell = ws.cell(row=1, column=current_col)
            cell.value = h
            cell.font = Font(bold=True)
            cell.fill = self.header_fill
            cell.border = self.thin_border
            current_col += 1

    def _write_new_sheet(self, ws, plans_data, results):
        self._write_headers(ws)

        # 样式
        font_normal = Font(name="微软雅黑", size=9)
        align = Alignment(vertical="center", wrap_text=False)

        for i, (plan, res) in enumerate(zip(plans_data, results)):
            row_idx = i + 2

            # 1. 写入基础列 (保留的 B-T)
            current_col = 1
            for col_idx in KEEP_COL_INDICES:
                cell = ws.cell(row=row_idx, column=current_col)
                val = plan['base_data'].get(col_idx)
                cell.value = val
                cell.font = font_normal
                cell.border = self.thin_border
                current_col += 1

            # 2. 写入计算结果列
            # 齐套率
            c_rate = ws.cell(row=row_idx, column=current_col)
            c_rate.value = res['rate']
            c_rate.number_format = '0%'
            c_rate.border = self.thin_border

            # 可产数量
            c_qty = ws.cell(row=row_idx, column=current_col + 1)
            c_qty.value = res['achievable']
            c_qty.border = self.thin_border

            # 缺料信息
            c_info = ws.cell(row=row_idx, column=current_col + 2)
            c_info.value = res['shortage_str']
            c_info.alignment = Alignment(wrap_text=True, vertical="center")  # 缺料信息自动换行
            c_info.border = self.thin_border

            # 3. 颜色标记 (整行标红或标绿)
            fill = self.red_fill if res['is_short'] else self.green_fill
            # 给新增加的3列上色，或者全行上色？
            # 需求通常是高亮显示状态。这里给后三列上色。
            c_rate.fill = fill
            c_qty.fill = fill
            c_info.fill = fill

        # 设置列宽
        ws.column_dimensions['A'].width = 15  # 假设A列是日期或单号
        # 最后一列缺料信息宽一点
        ws.column_dimensions[openpyxl.utils.get_column_letter(len(KEEP_COL_INDICES) + 3)].width = 50


if __name__ == "__main__":
    try:
        root = tk.Tk()
        app = DailyPlanAvailabilityApp(root)
        root.mainloop()
    except Exception as e:
        import tkinter.messagebox

        tkinter.messagebox.showerror("启动失败", str(e))