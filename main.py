# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.cell.cell import MergedCell # 核心：导入合并单元格判断类
import pyodbc
import traceback
import datetime
import copy
import math
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from tkcalendar import DateEntry

# ============== 1. 数据库查询逻辑 (锁定) ==============
def get_best_sql_driver():
    try:
        installed_drivers = [d for d in pyodbc.drivers()]
    except Exception: return "SQL Server"
    driver_preference = ["ODBC Driver 18 for SQL Server", "ODBC Driver 17 for SQL Server", "SQL Server"]
    for drv in driver_preference:
        if drv in installed_drivers: return drv
    return "SQL Server"

CURRENT_DRIVER = get_best_sql_driver()
DB_CONN_STRING = f"DRIVER={{{CURRENT_DRIVER}}};SERVER=192.168.0.117;DATABASE=FQD;UID=zhitan;PWD=Zt@forcome;TrustServerCertificate=yes;"

# ============== 2. 业务基础配置 ==============
ROW_IDX_DATA_START = 4   
COL_NAME_WORKSHOP = "车间"
COL_NAME_WO_TYPE = "单别"
COL_NAME_WO_NO = "工单单号"

class DailyPlanAvailabilityApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"每日排程齐套分析 v12.6 (A列强制显示版) - {CURRENT_DRIVER}")
        self.root.geometry("1150x750")

        self.red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")     
        self.green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")   
        self.yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  
        self.gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")    
        
        self.file_path = tk.StringVar()
        self.sheet_name = tk.StringVar()
        self.selected_workshop = tk.StringVar()
        self.is_date_range = tk.BooleanVar(value=False)
        self.date_column_map = {}
        self.col_map_main = {}
        self._create_widgets()

    def _create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        file_frame = ttk.LabelFrame(main_frame, text="1. 选择数据源", padding="5")
        file_frame.pack(fill=tk.X, pady=5)
        ttk.Entry(file_frame, textvariable=self.file_path, width=50).pack(side=tk.LEFT, padx=5)
        ttk.Button(file_frame, text="浏览...", command=self._select_file).pack(side=tk.LEFT, padx=5)
        self.sheet_combo = ttk.Combobox(file_frame, textvariable=self.sheet_name, state="disabled", width=15)
        self.sheet_combo.pack(side=tk.LEFT, padx=5)
        self.sheet_combo.bind("<<ComboboxSelected>>", self._on_sheet_selected)
        
        filter_frame = ttk.LabelFrame(main_frame, text="2. 分析参数", padding="10")
        filter_frame.pack(fill=tk.X, pady=5)
        date_frame = ttk.Frame(filter_frame); date_frame.pack(side=tk.LEFT, fill=tk.X)
        ttk.Checkbutton(date_frame, text="日期范围", variable=self.is_date_range, command=self._toggle_date_mode).pack(side=tk.LEFT, padx=(0, 10))
        self.date_start = DateEntry(date_frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy/mm/dd'); self.date_start.pack(side=tk.LEFT, padx=5)
        self.lbl_end = ttk.Label(date_frame, text="至"); self.date_end = DateEntry(date_frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy/mm/dd')
        self._toggle_date_mode()
        ttk.Label(filter_frame, text="车间:").pack(side=tk.LEFT, padx=(20, 5))
        self.workshop_combo = ttk.Combobox(filter_frame, textvariable=self.selected_workshop, state="disabled", width=15); self.workshop_combo.pack(side=tk.LEFT, padx=5)
        
        action_frame = ttk.LabelFrame(main_frame, text="3. 执行", padding="10"); action_frame.pack(fill=tk.X, pady=10)
        ttk.Button(action_frame, text="执行分析并另存副本 (A列标注)", command=self._run_analysis).pack(fill=tk.X, padx=100)
        self.log_text = tk.Text(main_frame, height=15, state="disabled", font=("Consolas", 9), bg="#F0F0F0"); self.log_text.pack(fill=tk.BOTH, expand=True, pady=5)

    def _toggle_date_mode(self):
        if self.is_date_range.get(): self.lbl_end.pack(side=tk.LEFT); self.date_end.pack(side=tk.LEFT, padx=5)
        else: self.lbl_end.pack_forget(); self.date_end.pack_forget()

    def _log(self, msg):
        self.log_text.config(state="normal"); self.log_text.insert(tk.END, f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {msg}\n"); self.log_text.see(tk.END); self.log_text.config(state="disabled"); self.root.update_idletasks()

    def _select_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if path:
            self.file_path.set(path)
            wb = openpyxl.load_workbook(path, read_only=True)
            self.sheet_combo['values'] = wb.sheetnames
            if wb.sheetnames: self.sheet_combo.current(0); self._on_sheet_selected(None)
            self.sheet_combo.config(state="readonly"); wb.close()

    def _on_sheet_selected(self, event):
        try:
            wb = openpyxl.load_workbook(self.file_path.get(), read_only=True, data_only=True)
            ws = wb[self.sheet_name.get()]; self.col_map_main = {}
            for r in [3, 2]: # v10.4 混合扫描
                for idx, cell in enumerate(ws[r], start=1):
                    val = str(cell.value).strip() if cell.value else ""
                    if val and val not in self.col_map_main: self.col_map_main[val] = idx
            self.date_column_map = {}
            for cell in ws[3]:
                dt = self._parse_excel_date(cell.value)
                if dt: self.date_column_map[dt] = cell.column
            col_ws_idx = self.col_map_main.get(COL_NAME_WORKSHOP)
            workshops = set()
            if col_ws_idx:
                for row in ws.iter_rows(min_row=ROW_IDX_DATA_START, min_col=col_ws_idx, max_col=col_ws_idx, values_only=True):
                    if row[0]: workshops.add(str(row[0]).strip())
            self.workshop_combo['values'] = ["全部车间"] + sorted(list(workshops)); self.workshop_combo.current(0); self.workshop_combo.config(state="readonly"); wb.close()
        except: self._log("读取表头出错")

    def _parse_excel_date(self, val):
        if val is None: return None
        try:
            if isinstance(val, (datetime.datetime, datetime.date)): return val.date()
            if isinstance(val, str):
                parts = val.strip().split('/')
                if len(parts) >= 2:
                    try: return datetime.datetime.strptime(val.strip(), "%Y/%m/%d").date()
                    except: return datetime.datetime.strptime(val.strip(), "%m/%d").replace(year=datetime.date.today().year).date()
            return None
        except: return None

    # ============== 3. ERP 查询与计算逻辑 (严格锁定) ==============
    def _fetch_erp_data(self, keys):
        if not keys: return {}
        conditions = [f"(TA.TA001='{t}' AND TA.TA002='{n}')" for t, n in keys]
        data = defaultdict(lambda: {'total': 0, 'bom': []})
        for i in range(0, len(conditions), 200):
            batch = conditions[i:i+200]
            sql = f"SELECT RTRIM(TA.TA001) t, RTRIM(TA.TA002) n, TA.TA015 total, RTRIM(TB.TB003) p, ISNULL(RTRIM(MB.MB002),'') name, ISNULL(RTRIM(MB.MB004),'') unit, TB.TB004 req, TB.TB005 iss FROM MOCTA TA INNER JOIN MOCTB TB ON TA.TA001=TB.TB001 AND TA.TA002=TB.TB002 LEFT JOIN INVMB MB ON TB.TB003=MB.MB001 WHERE {' OR '.join(batch)}"
            with pyodbc.connect(DB_CONN_STRING) as conn:
                df = pd.read_sql(sql, conn)
                for _, r in df.iterrows():
                    data[(r['t'], r['n'])]['total'] = float(r['total'])
                    data[(r['t'], r['n'])]['bom'].append({'part':r['p'],'name':r['name'],'unit':r['unit'],'req':float(r['req']),'iss':float(r['iss'])})
        return data

    def _fetch_inventory(self, parts):
        if not parts: return {}
        inv = {}
        parts_list = list(set(parts))
        for i in range(0, len(parts_list), 500):
            p_str = ",".join(f"'{p}'" for p in parts_list[i:i+500])
            sql = f"SELECT RTRIM(MC001) p, SUM(MC007) q FROM INVMC WHERE MC001 IN ({p_str}) GROUP BY MC001"
            with pyodbc.connect(DB_CONN_STRING) as conn:
                df = pd.read_sql(sql, conn); inv.update(pd.Series(df.q.values, index=df.p).to_dict())
        return inv

    # ============== 4. 核心处理流程 (修复合并单元格及隐藏 A 列问题) ==============
    def _run_analysis(self):
        start_dt = self.date_start.get_date(); end_dt = self.date_end.get_date() if self.is_date_range.get() else start_dt
        valid_dates = sorted([d for d in self.date_column_map if start_dt <= d <= end_dt])
        if not valid_dates: messagebox.showwarning("提示", "日期范围无产量数据列"); return

        dt_str = start_dt.strftime('%Y%m%d') if start_dt == end_dt else f"{start_dt.strftime('%Y%m%d')}-{end_dt.strftime('%Y%m%d')}"
        save_path = filedialog.asksaveasfilename(initialfile=f"{dt_str}齐套信息.xlsx", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not save_path: return

        try:
            self._log("第一步：预读排产计划...")
            plan_data_rows, all_wo_keys = [], set()
            c_type, c_no, c_ws = self.col_map_main.get(COL_NAME_WO_TYPE, 5), self.col_map_main.get(COL_NAME_WO_NO, 6), self.col_map_main.get(COL_NAME_WORKSHOP)
            target_ws = self.selected_workshop.get()
            
            wb_fast = openpyxl.load_workbook(self.file_path.get(), read_only=True, data_only=True)
            ws_fast = wb_fast[self.sheet_name.get()]
            
            for d in valid_dates:
                col_idx = self.date_column_map[d]
                for ridx, row_cells in enumerate(ws_fast.iter_rows(min_row=ROW_IDX_DATA_START), start=ROW_IDX_DATA_START):
                    qty = row_cells[col_idx-1].value
                    if isinstance(qty, (int, float)) and qty > 0:
                        if c_ws and target_ws != "全部车间" and str(row_cells[c_ws-1].value).strip() != target_ws: continue
                        wt, wn = row_cells[c_type-1].value, row_cells[c_no-1].value
                        if wt and wn:
                            key = (str(wt).strip(), str(wn).strip()); all_wo_keys.add(key)
                            plan_data_rows.append((ridx, d, int(round(float(qty))), key))
            wb_fast.close()

            if not all_wo_keys: messagebox.showinfo("完成", "无产量记录"); return

            self._log("第二步：同步 ERP 与库存数据..."); static_wo_data = self._fetch_erp_data(list(all_wo_keys))
            inv = self._fetch_inventory([b['part'] for w in static_wo_data.values() for b in w['bom']])

            self._log("第三步：执行库存分析..."); running_inv, running_issued = copy.deepcopy(inv), defaultdict(float)
            for k, v in static_wo_data.items():
                for b in v['bom']: running_issued[(k[0], k[1], b['part'])] = b['iss']
            
            final_res, final_col = defaultdict(list), defaultdict(int)

            for ridx, d, plan_qty, key in plan_data_rows:
                info = static_wo_data.get(key)
                if not info: continue
                max_erp = 999999
                for b in info['bom']:
                    u = b['req']/info['total'] if info['total']>0 else 0
                    if u > 0:
                        rem = max(0, b['req'] - running_issued.get((key[0], key[1], b['part']), 0))
                        max_erp = min(max_erp, int(rem // u))
                
                net = min(plan_qty, max_erp); excess = plan_qty - net
                min_rate, stock_can_do, shorts = 1.0, 999999, []
                for b in info['bom']:
                    u = b['req']/info['total'] if info['total']>0 else 0
                    if u > 0:
                        stk = max(0, running_inv.get(b['part'], 0))
                        if net > 0:
                            rate = stk / (net * u)
                            if rate < min_rate: min_rate = rate
                        stock_can_do = min(stock_can_do, int(stk // u))
                        if stk < (net * u) - 0.001: 
                            shorts.append(f"{b['name']}({b['part']})缺{(net*u)-stk:g}{b['unit']}")
                        running_inv[b['part']] = running_inv.get(b['part'], 0) - (plan_qty * u)
                        running_issued[(key[0], key[1], b['part'])] = running_issued.get((key[0], key[1], b['part']), 0) + (plan_qty * u)

                msg = f"齐套率为{min_rate:.0%}；可产数量为{min(net, stock_can_do)}个；工单净需求量为{net}个；超出工单的数量为{excess}个；此工单的缺料信息：{','.join(shorts) if shorts else '无'}"
                if len(valid_dates) > 1: msg = f"[{d.strftime('%m-%d')}] " + msg
                final_res[ridx].append(msg)
                prio = 1
                if net == 0 and excess > 0: prio = 2
                elif min_rate < 0.999: prio = 4
                elif excess > 0: prio = 3
                if prio > final_col[ridx]: final_col[ridx] = prio

            self._log("第四步：回写标注并强制显示 A 列...")
            wb_write = openpyxl.load_workbook(self.file_path.get())
            target_sn = self.sheet_name.get()
            for sn in wb_write.sheetnames:
                if sn != target_sn: del wb_write[sn]
            
            ws = wb_write[target_sn]
            
            # --- 关键修复：修改 A2 (主单元格)，跳过 A3 (从属合并格) ---
            # 这样即修改了标题，又避开了 MergedCell 只读报错
            cell_a2 = ws.cell(row=2, column=1)
            if not isinstance(cell_a2, MergedCell):
                cell_a2.value = "齐套信息"
                cell_a2.font = Font(bold=True)
                cell_a2.alignment = Alignment(horizontal="center", vertical="center")
            
            # 强制显示并拉宽 A 列
            ws.column_dimensions['A'].hidden = False
            ws.column_dimensions['A'].width = 85
            
            for ridx, msgs in final_res.items():
                cell = ws.cell(row=ridx, column=1)
                # 数据行如果存在合并，同样需要保护
                if not isinstance(cell, MergedCell):
                    cell.value = "\n".join(msgs)
                    cell.alignment = Alignment(wrapText=True, vertical="center")
                    cell.font = Font(size=9)
                    cp = final_col[ridx]
                    if cp == 4: cell.fill = self.red_fill
                    elif cp == 3: cell.fill = self.yellow_fill
                    elif cp == 2: cell.fill = self.gray_fill
                    elif cp == 1: cell.fill = self.green_fill

            wb_write.save(save_path)
            self._log("分析完成！副本已成功导出并显示。"); messagebox.showinfo("成功", "齐套信息已标注并强制显示。")
            
        except Exception as e:
            traceback.print_exc(); self._log(f"程序报错: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk(); app = DailyPlanAvailabilityApp(root); root.mainloop()
